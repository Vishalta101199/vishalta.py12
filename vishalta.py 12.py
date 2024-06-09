import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pytest

driver = webdriver.Chrome()
wait = WebDriverWait(driver, 10)

test_data = openpyxl.load_workbook('test_data.xlsx')
sheet = test_data.active

@pytest.mark.parametrize(['username', 'password'], [(sheet.cell(row=2, column=2).value, sheet.cell(row=2, column=3).value),
                                                  (sheet.cell(row=3, column=2).value, sheet.cell(row=3, column=3).value),
                                                  (sheet.cell(row=4, column=2).value, sheet.cell(row=4, column=3).value),
                                                  (sheet.cell(row=5, column=2).value, sheet.cell(row=5, column=3).value),
                                                  (sheet.cell(row=6, column=2).value, sheet.cell(row=6, column=3).value)])
def test_login(username, password):
    driver.get('https://opensource-demo.orangehrmlive.com/web/index.php/auth/login')
    
    username_field = wait.until(EC.presence_of_element_located((By.NAME, 'txtUsername')))
    username_field.send_keys(username)
    
    password_field = driver.find_element(By.NAME, 'txtPassword')
    password_field.send_keys(password)
    
    login_button = driver.find_element(By.ID, 'btnLogin')
    login_button.click()
    
    try:
        welcome_message = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div/div[1]/h1')))
        sheet.cell(row=2, column=4, value=time.strftime('%Y-%m-%d'))
        sheet.cell(row=2, column=5, value=time.strftime('%H:%M:%S'))
        sheet.cell(row=2, column=6, value='Tester1')
        sheet.cell(row=2, column=7, value='Passed')
    except:
        sheet.cell(row=2, column=4, value=time.strftime('%Y-%m-%d'))
        sheet.cell(row=2, column=5, value=time.strftime('%H:%M:%S'))
        sheet.cell(row=2, column=6, value='Tester1')
        sheet.cell(row=2, column=7, value='Failed')
    
    test_data.save('test_data.xlsx')

if _name_ == '_main_':
    pytest.main()